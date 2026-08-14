"""Reformat Planters_in_Malls.xlsx (3 sheets + 25 photos) into FAE master format."""
import os
import re
import zipfile
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from fmt import render, remap, merge_runs, place_image, finalize, HDR_FILL

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'
SRC = U + '43ae292f-Planters_in_Malls.xlsx'
IMG = 'planter_imgs'
os.makedirs(IMG, exist_ok=True)

# ── extract media, and map (sheet, anchor) -> file, so photos keep their meaning ──
z = zipfile.ZipFile(SRC)
for n in z.namelist():
    if n.startswith('xl/media/'):
        open(os.path.join(IMG, os.path.basename(n)), 'wb').write(z.read(n))

anchors = {}   # drawing index -> list of (cell, image file)
for dn in sorted(n for n in z.namelist() if re.match(r'xl/drawings/drawing\d+\.xml$', n)):
    idx = int(dn.split('drawing')[-1].split('.')[0])
    rels = z.read(f'xl/drawings/_rels/drawing{idx}.xml.rels').decode()
    relmap = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
    body = z.read(dn).decode()
    out = []
    for a in re.findall(r'<xdr:(?:two|one)CellAnchor[\s\S]*?</xdr:(?:two|one)CellAnchor>', body):
        f = re.search(r'<xdr:from><xdr:col>(\d+)</xdr:col>.*?<xdr:row>(\d+)</xdr:row>', a, re.S)
        rid = re.search(r'r:embed="(rId\d+)"', a)
        if f and rid:
            out.append((f'{get_column_letter(int(f.group(1)) + 1)}{int(f.group(2)) + 1}',
                        os.path.basename(relmap[rid.group(1)])))
    anchors[idx] = out

srcwb = openpyxl.load_workbook(SRC)
wb = openpyxl.Workbook()
wb.remove(wb.active)


def place(ws, cell, imgfile, w=128, h=140):
    im = XLImage(os.path.join(IMG, imgfile))
    im.width, im.height = w, h
    ws.add_image(im, cell)


# ───────────────── 1. Cinema & Ice Rink lobbies ─────────────────
src = srcwb['RC & IR Lobbies']
cols = [('S.No', None, 10.45), ('Location', None, 34.45),
        ('Small Pots ', None, 18.54), ('Remarks', None, 26.45)]
colmap = {'A': 'A', 'C': 'B', 'D': 'C'}
rowmap = {r: r for r in range(1, 30)}
grid = [[src.cell(row=r, column=1).value, src.cell(row=r, column=3).value,
         remap(src.cell(row=r, column=4).value, colmap, rowmap), None]
        for r in range(3, 24)]
ws = wb.create_sheet('RC & IR Lobbies')
render(ws, cols, grid, style='compact', freeze='A3', photo_row=True, photo_cols={3},
       total_rows={20}, label_span=2)
place(ws, 'C2', dict(anchors[1])['B3'])

# ───────────────── 2. Fashion parking lobbies ─────────────────
src = srcwb['FP Lobbies']
cols = [('S.No', None, 10.45), ('Location', None, 40.45),
        ('Small Pots ', None, 18.54), ('Medium Pots', None, 18.54),
        ('Remarks', None, 26.45)]
colmap = {'A': 'A', 'C': 'B', 'D': 'C', 'E': 'D', 'F': 'E'}
rowmap = {r: r - 1 for r in range(1, 30)}
grid = []
for r in list(range(4, 16)) + [19, 20]:
    grid.append([src.cell(row=r, column=1).value, src.cell(row=r, column=3).value]
                + [remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in (4, 5, 6)])
ws = wb.create_sheet('FP Lobbies')
render(ws, cols, grid, style='compact', freeze='A3', photo_row=True, photo_cols={3, 4},
       total_rows={11}, label_span=2)
place(ws, 'C2', dict(anchors[2])['B4'])

# ───────────────── 3. Mall planters ─────────────────
src = srcwb['Mall']
sizes = [src.cell(row=2, column=c).value for c in range(7, 11)]      # G2..J2
cols = ([('Location', None, 18.45), ('Sub location', None, 38.45)]
        + [('Planter Details', s, 18.54) for s in sizes]
        + [('Remarks', None, 26.45)])
colmap = {'E': 'A', 'F': 'B', 'G': 'C', 'H': 'D', 'I': 'E', 'J': 'F'}
rowmap = {r: r for r in range(1, 50)}

locs, loc = [], None
for r in range(3, 46):
    v = src.cell(row=r, column=5).value
    if v is not None:
        loc = v
    locs.append(loc)

grid = []
for i, r in enumerate(range(3, 46)):
    grid.append([locs[i], src.cell(row=r, column=6).value]
                + [remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in range(7, 11)]
                + [None])
grid.append([remap(src.cell(row=46, column=c).value, colmap, rowmap) for c in range(5, 11)] + [None])
grid[-1][0] = 'Total'

ws = wb.create_sheet('Mall Planters')
render(ws, cols, grid, style='compact', freeze='C3', photo_row=True,
       photo_cols={3, 4, 5, 6}, total_rows={len(grid) - 1}, label_span=2)
merge_runs(ws, 1, 3, locs)
# representative photo per size, taken from the source's own size legend panels
for cell, key in (('C2', 'A5'), ('D2', 'B5'), ('E2', 'C5'), ('F2', 'D5')):
    place(ws, cell, dict(anchors[3])[key])

# ───────────────── 4. Photo reference tab (every source photo, labelled) ─────────────────
SIZE_OF = {'A': 'Extra large', 'B': 'Big', 'C': 'Medum', 'D': 'Small'}
BLOCK = {5: 'GF block', 12: 'GF block', 19: 'FF block', 34: 'SF block',
         3: 'header', 4: 'header', 45: 'MLB'}
sheet_of = {1: 'RC & IR Lobbies', 2: 'FP Lobbies', 3: 'Mall'}

grid, ordered = [], []
for didx in sorted(anchors):
    for cell, img in anchors[didx]:
        col, rownum = re.match(r'([A-Z]+)(\d+)', cell).groups()
        label = (f'{SIZE_OF.get(col, col)} - {BLOCK.get(int(rownum), "row " + rownum)}'
                 if didx == 3 else f'cell {cell}')
        grid.append([sheet_of[didx], label, None, None])
        ordered.append(img)

ws = wb.create_sheet('Planter Photo Reference')
cols = [('Source sheet', None, 22.45), ('Shown at', None, 26.45),
        ('Photo', None, 22.45), ('Remarks', None, 24.45)]
render(ws, cols, grid, style='compact', freeze='C3', data_height=115.0)
for i, img in enumerate(ordered):
    place_image(ws, 3, 3 + i, os.path.join(IMG, img), w=128, h=145)

finalize(wb)
wb.save(OUT + 'Planters_in_Malls_Details_2026.xlsx')
print('wrote Planters_in_Malls_Details_2026.xlsx', wb.sheetnames)
