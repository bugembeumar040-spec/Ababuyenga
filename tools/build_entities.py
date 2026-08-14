"""Split the consolidated reports into one workbook per entity.

Zabeel, FV, China Town, TDM and FP each get their own register, one sheet per
asset type - the same shape as the FAE master. FAE itself is excluded; it
already has its own workbook.
"""
import io
from copy import copy, deepcopy
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from fmt import remap, finalize

R = '/home/user/Ababuyenga/reports/'
SRC = R + 'FAE_TDM_Assets_Details_2026_Combined.xlsx'
EMU = 9525
comb = openpyxl.load_workbook(SRC)

# openpyxl hands out a single stream per embedded image and closes it once the
# workbook is saved, so snapshot the bytes before writing anything.
PHOTOS = {}
for _ws in comb.worksheets:
    shots = []
    for _img in _ws._images:
        ref = _img.ref
        if not isinstance(ref, (bytes, bytearray)):
            ref.seek(0)
            ref = ref.read()
        shots.append((deepcopy(_img.anchor), bytes(ref)))
    PHOTOS[_ws.title] = shots


def _copy_cell(c, n):
    n.value = c.value
    if c.has_style:
        n.font, n.fill = copy(c.font), copy(c.fill)
        n.border, n.alignment = copy(c.border), copy(c.alignment)
        n.number_format = c.number_format
        n.protection = copy(c.protection)


def _copy_images(src, tgt, rowmap=None):
    for anchor, data in PHOTOS.get(src.title, ()):
        new = XLImage(io.BytesIO(data))
        new.anchor = deepcopy(anchor)
        if rowmap is not None:
            old = new.anchor._from.row + 1               # anchors are 0-based
            if old not in rowmap:
                continue
            new.anchor._from.row = rowmap[old] - 1
        ext = getattr(new.anchor, 'ext', None)
        if ext is not None:
            new.width, new.height = ext.cx / EMU, ext.cy / EMU
        tgt.add_image(new)


def copy_sheet(src, tgt):
    for k, d in src.column_dimensions.items():
        tgt.column_dimensions[k].width = d.width
    for k, d in src.row_dimensions.items():
        tgt.row_dimensions[k].height = d.height
    for row in src.iter_rows():
        for c in row:
            _copy_cell(c, tgt.cell(row=c.row, column=c.column))
    for rng in src.merged_cells.ranges:
        tgt.merge_cells(str(rng))
    tgt.freeze_panes = src.freeze_panes
    tgt.sheet_view.zoomScale = src.sheet_view.zoomScale
    _copy_images(src, tgt)


def subset_sheet(src, tgt, keep, header_rows=2):
    """Copy the header plus only `keep` (source row numbers), renumbering them
    from header_rows+1 down and remapping formulas to their new positions."""
    for k, d in src.column_dimensions.items():
        tgt.column_dimensions[k].width = d.width

    rowmap = {r: r for r in range(1, header_rows + 1)}
    for i, r in enumerate(keep):
        rowmap[r] = header_rows + 1 + i
    colmap = {}                                   # columns are unchanged

    for r in range(1, header_rows + 1):           # header
        tgt.row_dimensions[r].height = src.row_dimensions[r].height
        for c in src[r]:
            _copy_cell(c, tgt.cell(row=r, column=c.column))
    for i, r in enumerate(keep):                  # data
        nr = header_rows + 1 + i
        tgt.row_dimensions[nr].height = src.row_dimensions[r].height
        for c in src[r]:
            n = tgt.cell(row=nr, column=c.column)
            _copy_cell(c, n)
            n.value = remap(c.value, colmap, rowmap)

    for rng in src.merged_cells.ranges:
        rows = list(range(rng.min_row, rng.max_row + 1))
        if all(x in rowmap for x in rows):
            new = [rowmap[x] for x in rows]
            if new == list(range(min(new), max(new) + 1)):     # still contiguous
                tgt.merge_cells(start_row=min(new), start_column=rng.min_col,
                                end_row=max(new), end_column=rng.max_col)
    tgt.freeze_panes = src.freeze_panes
    tgt.sheet_view.zoomScale = src.sheet_view.zoomScale
    _copy_images(src, tgt, rowmap)


def build(filename, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for spec in sheets:
        if len(spec) == 2:
            srcname, newname = spec
            copy_sheet(comb[srcname], wb.create_sheet(newname))
        else:
            srcname, newname, keep, hdr = spec
            subset_sheet(comb[srcname], wb.create_sheet(newname), keep, hdr)
    finalize(wb)
    wb.save(R + filename)
    print(f'{filename:42} {len(wb.sheetnames)} tabs  '
          f'{sum(len(w._images) for w in wb.worksheets)} photos  {wb.sheetnames}')


# ───────────────────────────── Zabeel ─────────────────────────────
build('Zabeel_Assets_Details_2026.xlsx', [
    ('Zabeel Washroom Details', 'Zabeel Washroom Details'),
    ('Zabeel Handicap Washroom', 'Zabeel Handicap Washroom'),
    ('Baby Room Details', 'Zabeel Baby Room', list(range(45, 49)), 2),
    ('Washroom Assets Summery', 'Zabeel Assets Summery', [5, 10], 2),
])

# ─────────────────────────── Fountain Views ───────────────────────
build('FV_Assets_Details_2026.xlsx', [
    ('FV Washroom Details', 'FV Washroom Details'),
    ('FV Handicap Washroom', 'FV Handicap Washroom'),
    ('Baby Room Details', 'FV Baby Room', list(range(56, 79)), 2),
    ('Store Inventory', 'FV Store Inventory', [5, 6], 2),
    ('Washroom Assets Summery', 'FV Assets Summery', [12], 2),
])

# ───────────────────────────── China Town ─────────────────────────
build('China_Town_Assets_Details_2026.xlsx', [
    ('CT Washroom Details', 'CT Washroom Details'),
    ('CT Handicap Washroom', 'CT Handicap Washroom'),
    ('Baby Room Details', 'CT Baby Room', list(range(49, 56)), 2),
    ('Washroom Assets Summery', 'CT Assets Summery', [13], 2),
])

# ────────────────────────── Fashion Parking ───────────────────────
photo_rows = [r for r in range(3, comb['Planter Photo Reference'].max_row + 1)
              if comb['Planter Photo Reference'].cell(row=r, column=1).value == 'FP Lobbies']
build('FP_Assets_Details_2026.xlsx', [
    ('FP Lobbies', 'FP Planters'),
    ('Planter Photo Reference', 'FP Planter Photos', photo_rows, 2),
    ('Common Area Bins Details', 'FP Common Area Bins', [11], 2),
    ('External Staff Washroom', 'FP Staff Washroom', [12], 2),
])

# ───────────────────────────── TDM ────────────────────────────────
build('TDM_Assets_Details_2026.xlsx', [
    # washrooms
    ('Mall Washroom Deatils', 'Mall Washroom Deatils'),
    ('Staff Washroom Deatils', 'Staff Washroom Deatils'),
    ('External Staff Washroom', 'External Staff Washroom'),
    ('Mall Handicap Washroom', 'Mall Handicap Washroom'),
    ('Baby Room Details', 'Mall Baby Room', list(range(3, 26)), 2),
    ('Washroom Assets Summery', 'TDM Assets Summery', [3, 6, 9, 11], 2),
    # TDM detailed register
    ('TDM Washroom (Detailed)', 'TDM Washroom (Detailed)'),
    ('TDM Staff Washroom (Detail)', 'TDM Staff Washroom (Detail)'),
    ('TDM Baby Room (Detailed)', 'TDM Baby Room (Detailed)'),
    ('TDM Handicap Washroom', 'TDM Handicap Washroom'),
    ('TDM Washroom Summery', 'TDM Washroom Summery'),
    # bins
    ('Common Area Bins Details', 'Common Area Bins Details'),
    ('Steel Bins Details 2025', 'Steel Bins Details 2025'),
    ('SF FC Bin Stations', 'SF FC Bin Stations'),
    ('LG FC Bin Stations', 'LG FC Bin Stations'),
    ('SF FC Bin Spares', 'SF FC Bin Spares'),
    ('LG FC Bin Spares', 'LG FC Bin Spares'),
    # planters
    ('RC & IR Lobbies', 'RC & IR Lobbies'),
    ('Mall Planters', 'Mall Planters'),
    ('Planter Photo Reference', 'Planter Photo Reference'),
    # food court & furniture
    ('LG Food Court Furniture', 'LG Food Court Furniture'),
    ('LG Sofa Measurements', 'LG Sofa Measurements'),
    ('Mall External Dining', 'Mall External Dining'),
    ('Store Inventory', 'SFFC Store Inventory', [3, 4], 2),
    ('SF FC Assets Tracker', 'SF FC Assets Tracker'),
    ('SF FC External', 'SF FC External'),
    ('SFFC Chairs In Use', 'SFFC Chairs In Use'),
    ('SFFC Tables In Use', 'SFFC Tables In Use'),
    ('SFFC Leather Sofa', 'SFFC Leather Sofa'),
    ('SF FC Summery', 'SF FC Summery'),
])
