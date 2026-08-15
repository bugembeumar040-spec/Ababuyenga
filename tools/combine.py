"""Combine all ten reformatted reports into a single workbook, with an Index tab.

Sheets are copied cell-by-cell with their styles, merges, dimensions, freeze
panes and images, so every tab is identical to its standalone version.
"""
from copy import copy, deepcopy
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter, column_index_from_string
from fmt import HDR_FILL, TOT_FILL, finalize

R = '/home/user/Ababuyenga/reports/'
OUT = R + 'Malls_Assets_Details_2026_Combined.xlsx'
EMU = 9525

# (workbook, source sheet, sheet name in the combined file, section, original source file)
PLAN = [
    # ── Mall side register, built on the master's own template ──
    ('Mall_Side_Assets_Details_2026.xlsx', 'Mall Side Washroom Deatils', None, 'Mall side register', 'Mall_side__Assets_Details_2026.xlsx'),
    ('Mall_Side_Assets_Details_2026.xlsx', 'Mall Side Prayer Room', None, 'Mall side register', 'Mall_side__Assets_Details_2026.xlsx'),
    ('Mall_Side_Assets_Details_2026.xlsx', 'Mall Side Common Area', None, 'Mall side register', 'Mall_side__Assets_Details_2026.xlsx'),
    # ── washrooms, cross-property ───────────────────────────────────────────
    ('Washroom_Summary_Details_2026.xlsx', 'Mall Washroom Deatils', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'Staff Washroom Deatils', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'External Staff Washroom', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'Mall Handicap Washroom', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'Zabeel Washroom Details', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'Zabeel Handicap Washroom', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'CT Washroom Details', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'CT Handicap Washroom', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'FV Washroom Details', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'FV Handicap Washroom', None, 'Washrooms', 'Washroom_Summery.xlsx'),
    ('Washroom_Summary_Details_2026.xlsx', 'Baby Room Details', None, 'Washrooms', 'Washroom_Summery.xlsx',
     dict(keep=list(range(3, 26)) + list(range(45, 79)))),
    ('Washroom_Summary_Details_2026.xlsx', 'Washroom Assets Summery', None, 'Washrooms', 'Washroom_Summery.xlsx',
     dict(keep=[3, 5, 6, 9, 10, 11, 12, 13, 15],
          retotal={11: ('B', 'G', 'SUM({c}3:{c}10)')})),
    # ── washrooms, TDM detailed register (renamed: names clash with the above)
    ('TDM_Mall_Washroom_Details_2026.xlsx', 'Mall Washroom Deatils', 'TDM Washroom (Detailed)', 'Washrooms - TDM detail', 'Mall_Washroom.xlsx'),
    ('TDM_Mall_Washroom_Details_2026.xlsx', 'Staff Washroom Deatils', 'TDM Staff Washroom (Detail)', 'Washrooms - TDM detail', 'Mall_Washroom.xlsx'),
    ('TDM_Mall_Washroom_Details_2026.xlsx', 'Baby Room Details', 'TDM Baby Room (Detailed)', 'Washrooms - TDM detail', 'Mall_Washroom.xlsx'),
    ('TDM_Mall_Washroom_Details_2026.xlsx', 'Handicap Washroom Details', 'TDM Handicap Washroom', 'Washrooms - TDM detail', 'Mall_Washroom.xlsx'),
    ('TDM_Mall_Washroom_Details_2026.xlsx', 'Washroom Assets Summery', 'TDM Washroom Summery', 'Washrooms - TDM detail', 'Mall_Washroom.xlsx',
     dict(keep=[3, 5, 6, 9, 10, 11, 12, 13, 14, 15, 16, 18, 19, 20],
          retotal={9: ('B', 'G', 'SUM({c}3:{c}8)'), 16: ('B', 'B', 'SUM({c}13:{c}15)')})),
    # ── bins ────────────────────────────────────────────────────────────────
    ('TDM_Common_Area_Bins_Details_2026.xlsx', 'Common Area Bins Details', None, 'Bins', 'Common_area__Bin.xlsx'),
    ('TDM_Steel_Bins_Details_2026.xlsx', 'Steel Bins Details 2025', None, 'Bins', 'Steel_Bins.xlsx'),
    ('Bin_Stations_in_Food_Courts_Details_2026.xlsx', 'SF FC Bin Stations', None, 'Bins', 'Bin_Stations_in_Food_Courts.xlsx'),
    ('Bin_Stations_in_Food_Courts_Details_2026.xlsx', 'LG FC Bin Stations', None, 'Bins', 'Bin_Stations_in_Food_Courts.xlsx'),
    ('Bin_Stations_in_Food_Courts_Details_2026.xlsx', 'SF FC Bin Spares', None, 'Bins', 'Bin_Stations_in_Food_Courts.xlsx'),
    ('Bin_Stations_in_Food_Courts_Details_2026.xlsx', 'LG FC Bin Spares', None, 'Bins', 'Bin_Stations_in_Food_Courts.xlsx'),
    # ── planters ────────────────────────────────────────────────────────────
    ('Planters_in_Malls_Details_2026.xlsx', 'RC & IR Lobbies', None, 'Planters', 'Planters_in_Malls.xlsx'),
    ('Planters_in_Malls_Details_2026.xlsx', 'FP Lobbies', None, 'Planters', 'Planters_in_Malls.xlsx'),
    ('Planters_in_Malls_Details_2026.xlsx', 'Mall Planters', None, 'Planters', 'Planters_in_Malls.xlsx'),
    ('Planters_in_Malls_Details_2026.xlsx', 'Planter Photo Reference', None, 'Planters', 'Planters_in_Malls.xlsx'),
    # ── food court & furniture ──────────────────────────────────────────────
    ('LG_FC_Furniture_Details_2026.xlsx', 'LG Food Court Furniture', None, 'Food court & furniture', 'LG_FC.xlsx'),
    ('LG_FC_Furniture_Details_2026.xlsx', 'LG Sofa Measurements', None, 'Food court & furniture', 'LG_FC.xlsx'),
    ('New_External_Dining_Furniture_Details_2026.xlsx', 'Mall External Dining', None, 'Food court & furniture', 'New_External_dining_furniture_details.xlsx'),
    ('New_External_Dining_Furniture_Details_2026.xlsx', 'Store Inventory', None, 'Food court & furniture', 'New_External_dining_furniture_details.xlsx'),
    ('SF_FC_Assets_Inspection_Tracker_2026.xlsx', 'SF FC Assets Tracker', None, 'Food court & furniture', 'SF_FC_ASSETS_INSPECTION_TRACKER2026.xlsx'),
    ('SF_FC_Furniture_Details_2026.xlsx', 'SF FC External', None, 'Food court & furniture', 'SF_FC.xlsx'),
    ('SF_FC_Furniture_Details_2026.xlsx', 'SFFC Chairs In Use', None, 'Food court & furniture', 'SF_FC.xlsx'),
    ('SF_FC_Furniture_Details_2026.xlsx', 'SFFC Tables In Use', None, 'Food court & furniture', 'SF_FC.xlsx'),
    ('SF_FC_Furniture_Details_2026.xlsx', 'SFFC Leather Sofa', None, 'Food court & furniture', 'SF_FC.xlsx'),
    ('SF_FC_Furniture_Details_2026.xlsx', 'SF FC Summery', None, 'Food court & furniture', 'SF_FC.xlsx'),
]


def copy_sheet(src, tgt):
    for coord, dim in src.column_dimensions.items():
        d = tgt.column_dimensions[coord]
        d.width, d.hidden = dim.width, dim.hidden
        d.bestFit = False
    for idx, dim in src.row_dimensions.items():
        tgt.row_dimensions[idx].height = dim.height

    for row in src.iter_rows():
        for c in row:
            n = tgt.cell(row=c.row, column=c.column)
            n.value = c.value
            if c.has_style:
                # style *objects*, not _style: the index array points into the
                # source workbook's style table and is meaningless here
                n.font = copy(c.font)
                n.fill = copy(c.fill)
                n.border = copy(c.border)
                n.alignment = copy(c.alignment)
                n.number_format = c.number_format
                n.protection = copy(c.protection)
    for rng in src.merged_cells.ranges:
        tgt.merge_cells(str(rng))

    tgt.freeze_panes = src.freeze_panes
    tgt.sheet_view.zoomScale = src.sheet_view.zoomScale
    tgt.sheet_view.showGridLines = src.sheet_view.showGridLines

    for img in src._images:
        data = img.ref
        try:
            data.seek(0)
        except Exception:
            pass
        new = XLImage(data)
        new.anchor = deepcopy(img.anchor)
        ext = getattr(new.anchor, 'ext', None)
        if ext is not None:                 # keep the on-sheet size, not the native size
            new.width, new.height = ext.cx / EMU, ext.cy / EMU
        tgt.add_image(new)


wb = openpyxl.Workbook()
wb.remove(wb.active)
index = wb.create_sheet('Index')



def subset_sheet(src, tgt, keep, retotal=None):
    """Copy the two header rows plus only `keep` source rows, renumbering them
    from row 3 down and remapping formulas. Used to drop FAE rows from tabs that
    mix several properties; any total whose range spanned a dropped row is
    rewritten over the retained range via `retotal`."""
    from fmt import remap
    for k, d in src.column_dimensions.items():
        tgt.column_dimensions[k].width = d.width
    rowmap = {1: 1, 2: 2}
    for i, r in enumerate(keep):
        rowmap[r] = 3 + i
    for r in (1, 2):
        tgt.row_dimensions[r].height = src.row_dimensions[r].height
        for c in src[r]:
            n = tgt.cell(row=r, column=c.column)
            n.value = c.value
            if c.has_style:
                n.font, n.fill = copy(c.font), copy(c.fill)
                n.border, n.alignment = copy(c.border), copy(c.alignment)
                n.number_format = c.number_format
    for i, r in enumerate(keep):
        nr = 3 + i
        tgt.row_dimensions[nr].height = src.row_dimensions[r].height
        for c in src[r]:
            n = tgt.cell(row=nr, column=c.column)
            n.value = remap(c.value, {}, rowmap)
            if c.has_style:
                n.font, n.fill = copy(c.font), copy(c.fill)
                n.border, n.alignment = copy(c.border), copy(c.alignment)
                n.number_format = c.number_format
    for rng in src.merged_cells.ranges:
        rr = list(range(rng.min_row, rng.max_row + 1))
        if all(x in rowmap for x in rr):
            new = [rowmap[x] for x in rr]
            if new == list(range(min(new), max(new) + 1)):
                tgt.merge_cells(start_row=min(new), start_column=rng.min_col,
                                end_row=max(new), end_column=rng.max_col)
    for nr, (c1, c2, tpl) in (retotal or {}).items():
        for ci in range(column_index_from_string(c1), column_index_from_string(c2) + 1):
            col = get_column_letter(ci)
            tgt.cell(row=nr, column=ci).value = '=' + tpl.format(c=col)
    tgt.freeze_panes = src.freeze_panes
    tgt.sheet_view.zoomScale = src.sheet_view.zoomScale


cache, rows = {}, []
for entry in PLAN:
    f, sheet, rename, section, origin = entry[:5]
    opts = entry[5] if len(entry) > 5 else None
    if f not in cache:
        cache[f] = openpyxl.load_workbook(R + f)
    src = cache[f][sheet]
    name = rename or sheet
    tgt = wb.create_sheet(name)
    if opts:
        subset_sheet(src, tgt, opts['keep'], opts.get('retotal'))
        note = 'Filtered - out-of-scope property rows removed'
    else:
        copy_sheet(src, tgt)
        note = None
    rows.append((section, name, origin, tgt.max_row - 2, tgt.max_column,
                 len(tgt._images), note))

# ───────────────────────── Index tab ─────────────────────────
heads = ['#', 'Section', 'Tab', 'Original source file', 'Data rows', 'Columns', 'Photos',
         'Remarks']
widths = [6.45, 24.45, 32.45, 44.45, 12.45, 11.45, 10.45, 26.45]
for i, (h, w) in enumerate(zip(heads, widths), start=1):
    c = index.cell(row=1, column=i, value=h)
    c.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    index.column_dimensions[get_column_letter(i)].width = w
index.row_dimensions[1].height = 37.4
index.freeze_panes = 'A2'

prev = None
for n, (section, name, origin, nrows, ncols, nimg, note) in enumerate(rows, start=1):
    r = n + 1
    index.row_dimensions[r].height = 22.0
    vals = [n, section if section != prev else '', name, origin, nrows, ncols, nimg or '', note]
    prev = section
    for i, v in enumerate(vals, start=1):
        c = index.cell(row=r, column=i, value=v)
        c.font = Font(name='Calibri', size=11, color='000000')
        c.alignment = Alignment(horizontal='center' if i != 3 and i != 4 else 'left',
                                vertical='center')
    link = index.cell(row=r, column=3)
    # internal jump: Excel needs `location`, not a target URL
    link.hyperlink = Hyperlink(ref=link.coordinate, location=f"'{name}'!A1")
    link.font = Font(name='Calibri', size=11, color='0563C1', underline='single')

tot = index.cell(row=len(rows) + 2, column=1, value=f'{len(rows)} report tabs')
tot.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
tot.fill = HDR_FILL
tot.alignment = Alignment(horizontal='center', vertical='center')
index.merge_cells(start_row=len(rows) + 2, start_column=1, end_row=len(rows) + 2, end_column=4)
for i, col in enumerate((5, 6, 7)):
    c = index.cell(row=len(rows) + 2, column=col,
                   value=sum(r[3 + i] for r in rows) if col != 6 else '')
    c.font = Font(name='Calibri', size=12, bold=True, color='000000')
    c.fill = TOT_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')

finalize(wb)
wb.save(OUT)
print('wrote', OUT)
print('tabs:', len(wb.sheetnames), '| images:', sum(len(ws._images) for ws in wb.worksheets))
