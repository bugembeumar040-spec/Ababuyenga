"""Batch 2: reformat the food-court / furniture inventories into FAE master format.

Rule used throughout: keep the FULL contiguous source row span (first data row
through the last total row) so every remapped formula reference stays valid.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from fmt import render, remap, merge_runs, place_image, extract_media, finalize

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'
A = 18.45   # standard asset column width


def cells(src, r, c1, c2, colmap, rowmap):
    return [remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in range(c1, c2 + 1)]


# ═══════════════ 1. LG Food Court furniture ═══════════════
SRCF = U + 'c78f005a-LG_FC.xlsx'
anch, s2d = extract_media(SRCF, 'b2_lgfc')
src = openpyxl.load_workbook(SRCF)['LG Food Court Furniture ']
wb = openpyxl.Workbook(); wb.remove(wb.active)

assets = [src.cell(row=2, column=c).value for c in range(1, 9)]
cols = ([('Level', None, 14.45), ('Location', None, 26.45)]
        + [(a, None, A) for a in assets] + [('Remarks', None, 26.45)])
grid = [['LG', 'LG Food Court'] + [src.cell(row=4, column=c).value for c in range(1, 9)] + [None]]
ws = wb.create_sheet('LG Food Court Furniture')
render(ws, cols, grid, style='compact', freeze='C3', photo_row=True,
       photo_cols=set(range(3, 11)))
for cell, img in anch[s2d[1]]:
    if cell.endswith('3'):                       # row-3 photo strip in the source
        col = openpyxl.utils.column_index_from_string(cell[:-1])
        place_image(ws, col + 2, 2, f'b2_lgfc/{img}', w=118, h=132)

# sofa measurement block (rows 8-36): a genuine two-row banded header
sub = [src.cell(row=9, column=c).value for c in range(3, 7)]
cols = ([('Location', None, 26.45), ('Sofa', None, 18.45)]
        + [('Straight sofa', sub[0], A), ('Straight sofa', sub[1], A),
           ('L Shape', sub[2], A), ('L Shape', sub[3], A)]
        + [('Remarks', None, 26.45)])
colmap = {c: c for c in 'ABCDEF'}
rowmap = {r: r - 7 for r in range(1, 40)}
locs, cur = [], None
for r in range(10, 37):
    v = src.cell(row=r, column=1).value
    if v is not None:
        cur = v
    locs.append(cur)
grid = [[locs[i]] + cells(src, r, 2, 6, colmap, rowmap) + [None]
        for i, r in enumerate(range(10, 37))]
ws = wb.create_sheet('LG Sofa Measurements')
render(ws, cols, grid, style='compact', freeze='C3')
merge_runs(ws, 1, 3, locs)
finalize(wb)
wb.save(OUT + 'LG_FC_Furniture_Details_2026.xlsx')
print('LG_FC_Furniture_Details_2026.xlsx', wb.sheetnames)

# ═══════════════ 2. Bin stations in food courts ═══════════════
SRCF = U + 'bfd9eba2-Bin_Stations_in_Food_Courts.xlsx'
anch, s2d = extract_media(SRCF, 'b2_bins')
swb = openpyxl.load_workbook(SRCF)
wb = openpyxl.Workbook(); wb.remove(wb.active)


def bin_station(src, title, hdr1, hdr2, first, last, c0, drawing, img_src_col):
    """SN | SUB LOCATION | IMAGE | TYPES OF BIN[5] | REMARKS"""
    subs = [src.cell(row=hdr2, column=c).value for c in range(c0 + 3, c0 + 8)]
    cols = ([('SN', None, 10.45), ('SUB LOCATION', None, 28.45), ('IMAGE', None, 20.45)]
            + [('TYPES OF BIN', s, 13.45) for s in subs]
            + [('REMARKS', None, 20.45)])
    colmap = {get_column_letter(c0 + i): get_column_letter(1 + i) for i in range(9)}
    rowmap = {r: r - first + 3 for r in range(1, src.max_row + 3)}
    grid = [cells(src, r, c0, c0 + 8, colmap, rowmap) for r in range(first, last + 1)]
    ws = wb.create_sheet(title)
    render(ws, cols, grid, style='compact', freeze='D3', total_rows={len(grid) - 1},
           label_span=2, data_height=84.95)
    seen = {}
    for cell, img in anch[drawing]:
        c, r = cell[:len(cell) - len(cell.lstrip('ABCDEFGHIJ'))], int(cell.lstrip('ABCDEFGHIJ'))
        if c != img_src_col or not (first <= r <= last):
            continue
        n = seen.get(r, 0); seen[r] = n + 1
        place_image(ws, 3, r - first + 3, f'b2_bins/{img}', w=88, h=104, dx=2 + n * 92)
    return ws


bin_station(swb['SF FC'], 'SF FC Bin Stations', 5, 6, 7, 24, 1, s2d[1], 'C')
bin_station(swb['LG FC'], 'LG FC Bin Stations', 4, 5, 6, 11, 2, s2d[2], 'D')


def bin_spares(src, title, hdr, first, last, c0, drawing, img_src_col):
    heads = [src.cell(row=hdr, column=c).value for c in range(c0, c0 + 5)]
    cols = [(heads[0], None, 24.45), (heads[1], None, 20.45)] + \
           [(h, None, A) for h in heads[2:]] + [('Remarks', None, 24.45)]
    colmap = {get_column_letter(c0 + i): get_column_letter(1 + i) for i in range(5)}
    rowmap = {r: r - first + 3 for r in range(1, src.max_row + 3)}
    grid = [cells(src, r, c0, c0 + 4, colmap, rowmap) + [None] for r in range(first, last + 1)]
    ws = wb.create_sheet(title)
    render(ws, cols, grid, style='compact', freeze='C3', data_height=84.95)
    for cell, img in anch[drawing]:
        c, r = cell[:len(cell) - len(cell.lstrip('ABCDEFGHIJ'))], int(cell.lstrip('ABCDEFGHIJ'))
        if c == img_src_col and first <= r <= last:
            place_image(ws, 2, r - first + 3, f'b2_bins/{img}', w=110, h=104)


bin_spares(swb['SF FC'], 'SF FC Bin Spares', 26, 27, 28, 2, s2d[1], 'C')
bin_spares(swb['LG FC'], 'LG FC Bin Spares', 14, 15, 16, 2, s2d[2], 'C')
finalize(wb)
wb.save(OUT + 'Bin_Stations_in_Food_Courts_Details_2026.xlsx')
print('Bin_Stations_in_Food_Courts_Details_2026.xlsx', wb.sheetnames)

# ═══════════════ 3. New external dining furniture ═══════════════
SRCF = U + '054e9b21-New_External_dining_furniture_details.xlsx'
anch, s2d = extract_media(SRCF, 'b2_ned')
swb = openpyxl.load_workbook(SRCF)
wb = openpyxl.Workbook(); wb.remove(wb.active)

src = swb['MALL']
cols = [('SL ', None, 10.45), ('Location', None, 26.45), ('Material description ', None, 26.45),
        ('Assets Picture ', None, 22.45), ('Available Stock ', None, A), ('Remarks', None, 24.45)]
colmap = {c: c for c in 'ABCDE'}
rowmap = {r: r for r in range(1, 20)}
locs, cur = [], None
for r in range(3, 11):
    v = src.cell(row=r, column=2).value
    if v is not None:
        cur = v
    locs.append(cur)
grid = []
for i, r in enumerate(range(3, 14)):
    row = cells(src, r, 1, 5, colmap, rowmap) + [None]
    if i < len(locs):
        row[1] = locs[i]
    grid.append(row)
ws = wb.create_sheet('Mall External Dining')
render(ws, cols, grid, style='compact', freeze='C3', total_rows={9, 10},
       label_span=2, data_height=100.15)
merge_runs(ws, 2, 3, locs)
for cell, img in anch[s2d[1]]:
    if cell.startswith('D'):
        place_image(ws, 4, int(cell[1:]), f'b2_ned/{img}', w=120, h=118)

# store inventory: two stores in one source sheet -> a leading Store column
src = swb['Store Inventory']
cols = [('Store', None, 34.45), ('S/R No', None, 12.45), ('Items', None, 22.45),
        ('Qty', None, A), ('Remarks', None, 24.45)]
grid = []
for store, rows in ((src['A2'].value, (4, 5)), (src['A6'].value, (8, 9))):
    for r in rows:
        grid.append([store] + [src.cell(row=r, column=c).value for c in range(1, 5)])
ws = wb.create_sheet('Store Inventory')
render(ws, cols, grid, style='compact', freeze='C3')
merge_runs(ws, 1, 3, [g[0] for g in grid])
ws.merge_cells('C3:C4')          # source B4:B5 - 'Table' spans both rows
finalize(wb)
wb.save(OUT + 'New_External_Dining_Furniture_Details_2026.xlsx')
print('New_External_Dining_Furniture_Details_2026.xlsx', wb.sheetnames)

# ═══════════════ 4. SF FC weekly asset inspection tracker ═══════════════
SRCF = U + '6566d45d-SF_FC_ASSETS_INSPECTION_TRACKER2026.xlsx'
anch, s2d = extract_media(SRCF, 'b2_track')
src = openpyxl.load_workbook(SRCF)['SF FC ASSETS TRACKER']
wb = openpyxl.Workbook(); wb.remove(wb.active)

heads = [src.cell(row=2, column=c).value for c in range(1, 10)]
cols = ([('SL', None, 10.45), (heads[1], None, 30.45), (heads[2], None, 22.45),
         (heads[3], None, 14.45)] + [(h, None, A) for h in heads[4:8]]
        + [('Remarks', None, 24.45)])
grid = [[src.cell(row=r, column=c).value for c in range(1, 10)] for r in range(3, 26)]
ws = wb.create_sheet('SF FC Assets Tracker')
render(ws, cols, grid, style='compact', freeze='D3', data_height=86.25)
seen = {}
for cell, img in anch[s2d[1]]:
    if cell.startswith('C'):
        r = int(cell[1:]); n = seen.get(r, 0); seen[r] = n + 1
        place_image(ws, 3, r, f'b2_track/{img}', w=78, h=100, dx=2 + n * 82)
finalize(wb)
wb.save(OUT + 'SF_FC_Assets_Inspection_Tracker_2026.xlsx')
print('SF_FC_Assets_Inspection_Tracker_2026.xlsx', wb.sheetnames)
