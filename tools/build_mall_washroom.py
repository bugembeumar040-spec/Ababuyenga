"""Reformat Mall_Washroom.xlsx (5 sheets) into FAE master format."""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from fmt import render, remap, merge_runs, HDR_FILL, TOT_FILL

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'

srcwb = openpyxl.load_workbook(U + 'a62f1923-Mall_Washroom.xlsx')
wb = openpyxl.Workbook()
wb.remove(wb.active)


def wash_sheet(src, out_title, hdr_row, first, last, total_offsets, last_col):
    """Master layout: Area | Washroom no | Washroom (Gender) | assets... | Remarks"""
    assets = [src.cell(row=hdr_row, column=c).value for c in range(3, last_col + 1)]
    cols = ([('Area', None, 30.45), ('Washroom no', None, 14.45),
             ('Washroom (Gender)', None, 20.45)]
            + [(a, None, 16.45) for a in assets]
            + [('Remarks', None, 30.45)])

    # src A->A, src B->C, src C..-> D..
    colmap = {'A': 'A', 'B': 'C'}
    for c in range(3, last_col + 1):
        colmap[get_column_letter(c)] = get_column_letter(c + 1)
    off = 3 - first
    rowmap = {r: r + off for r in range(1, src.max_row + 2)}

    # resolve merged Area labels, then number washrooms within each Area block
    areas, area = [], None
    for r in range(first, last + 1):
        v = src.cell(row=r, column=1).value
        if v is not None:
            area = v
        areas.append(area)
    seq, prev, n = [], object(), 0
    for a in areas:
        n = 1 if a != prev else n + 1
        seq.append(n)
        prev = a

    grid = []
    for i, r in enumerate(range(first, last + 1)):
        grid.append([areas[i], seq[i], src.cell(row=r, column=2).value]
                    + [remap(src.cell(row=r, column=c).value, colmap, rowmap)
                       for c in range(3, last_col + 1)] + [None])
    for r in total_offsets:
        grid.append([remap(src.cell(row=r, column=1).value, colmap, rowmap), None,
                     remap(src.cell(row=r, column=2).value, colmap, rowmap)]
                    + [remap(src.cell(row=r, column=c).value, colmap, rowmap)
                       for c in range(3, last_col + 1)] + [None])

    ws = wb.create_sheet(out_title)
    totals = set(range(len(grid) - len(total_offsets), len(grid)))
    render(ws, cols, grid, style='washroom', freeze='D3',
           total_rows=totals, label_span=2)
    merge_runs(ws, 1, 3, areas)
    return ws


wash_sheet(srcwb[' Washroom'], 'Mall Washroom Deatils', 2, 4, 63, [64, 65, 66], 29)
wash_sheet(srcwb['Staff Washroom '], 'Staff Washroom Deatils', 2, 4, 14, [15, 16, 17], 18)
wash_sheet(srcwb['Baby Room'], 'Baby Room Details', 2, 3, 24, [25, 26, 27], 22)
hw = wash_sheet(srcwb['Handcup Washroom'], 'Handicap Washroom Details', 2, 4, 11, [12, 13, 14], 14)
# source merged C6:N6 and C9:N9 (single value spanning the asset columns)
hw.merge_cells('D5:O5')
hw.merge_cells('D8:O8')

# ───────────────────────── Summary sheet ─────────────────────────
src = srcwb['Summery ']
cols = [('Loaction ', None, 26.45)] + [(src.cell(row=3, column=c).value, None, 18.45)
                                       for c in range(5, 11)] + [('Remarks', None, 26.45)]
colmap = {get_column_letter(c): get_column_letter(c - 3) for c in range(4, 11)}
rowmap = {r: r - 1 for r in range(1, 30)}
grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in range(4, 11)] + [None]
        for r in range(4, 14)]

ws = wb.create_sheet('Washroom Assets Summery')
render(ws, cols, grid, style='compact', freeze='B3', total_rows={9}, label_span=1)

# second block: Pedal Bins at Washroom (source H20:I25)
ws.cell(row=15, column=1, value='Pedal Bins at Washroom').font = Font(
    name='Calibri', size=12, bold=True, color='FFFFFF')
ws.cell(row=15, column=1).fill = HDR_FILL
ws.cell(row=15, column=1).alignment = Alignment(horizontal='center', vertical='center')
for i, r in enumerate(range(21, 26)):
    rr = 16 + i
    ws.row_dimensions[rr].height = 25.0
    lbl = ws.cell(row=rr, column=1, value=src.cell(row=r, column=8).value)
    val = ws.cell(row=rr, column=2, value=remap(src.cell(row=r, column=9).value,
                                                {'I': 'B'}, {x: x - 5 for x in range(1, 30)}))
    lbl.font = Font(name='Calibri', size=12, color='000000')
    val.font = Font(name='Calibri', size=14, bold=True, color='000000')
    val.fill = TOT_FILL
    for c in (lbl, val):
        c.alignment = Alignment(horizontal='center', vertical='center')

wb.save(OUT + 'TDM_Mall_Washroom_Details_2026.xlsx')
print('wrote TDM_Mall_Washroom_Details_2026.xlsx', wb.sheetnames)
