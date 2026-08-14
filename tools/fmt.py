"""Shared formatter reproducing the FAE_Assets_Details_2026 master workbook style."""
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

HDR_HEX = '44546A'     # theme dk2 - header band
TOT_HEX = 'FFC000'     # theme accent4 - total row values
HDR_FILL = PatternFill('solid', fgColor=HDR_HEX)
TOT_FILL = PatternFill('solid', fgColor=TOT_HEX)

THIN = Side(style='thin')
MED = Side(style='medium')

STYLE = {
    # washroom / prayer-room tabs
    'washroom': dict(hdr_sz=16, data_sz=18, tot_lbl_sz=18, tot_val_sz=20,
                     tot_lbl_bold=False, h1=55.4, h2=44.15, hd=33.65),
    # common-area style tabs (bins, planters, summaries)
    'compact': dict(hdr_sz=12, data_sz=12, tot_lbl_sz=18, tot_val_sz=14,
                    tot_lbl_bold=True, h1=37.4, h2=37.4, hd=25.0),
}


def _border(left=THIN, right=THIN, top=THIN, bottom=THIN):
    return Border(left=left, right=right, top=top, bottom=bottom)


def render(ws, cols, grid, *, style='washroom', freeze='A3', photo_row=False,
           photo_cols=(), total_rows=(), label_span=None, first_row=None,
           data_height=None):
    """Lay out a sheet in master format.

    cols  : list of (group, sub, width). sub=None -> the column is flat and its
            header is vertically merged across rows 1-2 (master's cols A-H rule).
            Consecutive columns sharing a non-None group are merged in row 1.
    grid  : list of row-lists written from the first data row down.
    total_rows : indices into `grid` that are total rows (amber fill).
    label_span : number of leading columns the total label is merged across.
    """
    st = STYLE[style]
    ncol = len(cols)
    data_start = first_row if first_row else (3 if not photo_row else 3)

    # ---------- column widths ----------
    for i, (_g, _s, w) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- group boundaries ----------
    # a column starts a new group if its group label differs from the previous
    starts, ends = set(), set()
    for i, (g, s, _w) in enumerate(cols, start=1):
        prev = cols[i - 2] if i >= 2 else None
        nxt = cols[i] if i < ncol else None
        if s is None or prev is None or prev[0] != g or prev[1] is None:
            starts.add(i)
        if s is None or nxt is None or nxt[0] != g or nxt[1] is None:
            ends.add(i)

    def hb(i, top, bottom):
        return _border(left=MED if i in starts else THIN,
                       right=MED if i in ends else THIN,
                       top=top, bottom=bottom)

    # ---------- header ----------
    ws.row_dimensions[1].height = st['h1']
    ws.row_dimensions[2].height = st['h2'] if not photo_row else 107.5
    hdr_font = Font(name='Calibri', size=st['hdr_sz'], bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    i = 1
    while i <= ncol:
        g, s, _w = cols[i - 1]
        if s is None:
            # flat column -> vertical merge across the two header rows, unless
            # row 2 is reserved for this column's reference photo
            is_photo = photo_row and i in photo_cols
            if not is_photo:
                ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
            c = ws.cell(row=1, column=i, value=g)
            c.font, c.fill, c.alignment = hdr_font, HDR_FILL, align
            c.border = hb(i, MED, THIN if is_photo else MED)
            if not is_photo:
                c2 = ws.cell(row=2, column=i)
                c2.font, c2.fill, c2.alignment = hdr_font, HDR_FILL, align
                c2.border = hb(i, MED, MED)
            i += 1
        else:
            j = i
            while j < ncol and cols[j][0] == g and cols[j][1] is not None:
                j += 1
            if j > i:
                ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=j)
            for k in range(i, j + 1):
                c = ws.cell(row=1, column=k, value=g if k == i else None)
                c.font, c.fill, c.alignment = hdr_font, HDR_FILL, align
                c.border = hb(k, MED, THIN)
                c2 = ws.cell(row=2, column=k, value=cols[k - 1][1])
                c2.font, c2.fill, c2.alignment = hdr_font, HDR_FILL, align
                c2.border = hb(k, THIN, MED)
            i = j + 1

    # photo strip cells keep the header boundaries but carry no text
    if photo_row:
        for k in photo_cols:
            ws.cell(row=2, column=k).border = hb(k, THIN, MED)

    # ---------- data ----------
    data_font = Font(name='Calibri', size=st['data_sz'], color='000000')
    tot_lbl_font = Font(name='Calibri', size=st['tot_lbl_sz'],
                        bold=st['tot_lbl_bold'], color='FFFFFF')
    tot_val_font = Font(name='Calibri', size=st['tot_val_sz'], bold=True, color='000000')

    for r_off, row in enumerate(grid):
        r = data_start + r_off
        ws.row_dimensions[r].height = st['hd'] if data_height is None else data_height
        is_total = r_off in total_rows
        for k in range(1, ncol + 1):
            v = row[k - 1] if k - 1 < len(row) else None
            c = ws.cell(row=r, column=k, value=v)
            c.alignment = align
            c.border = hb(k, THIN, THIN)
            if is_total:
                if label_span and k <= label_span:
                    c.font, c.fill = tot_lbl_font, HDR_FILL
                else:
                    c.font, c.fill = tot_val_font, TOT_FILL
            else:
                c.font = data_font
        if is_total and label_span and label_span > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=label_span)

    if freeze:
        ws.freeze_panes = freeze
    return ws


def merge_runs(ws, col, first_row, values):
    """Vertically merge repeated values in an identity column (master's Area style)."""
    start = None
    for idx in range(len(values) + 1):
        cur = values[idx] if idx < len(values) else object()
        prev = values[idx - 1] if idx > 0 else None
        if idx > 0 and cur == prev and prev is not None:
            if start is None:
                start = idx - 1
        else:
            if start is not None and idx - 1 > start:
                ws.merge_cells(start_row=first_row + start, start_column=col,
                               end_row=first_row + idx - 1, end_column=col)
            start = None


_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')


def remap(formula, colmap, rowmap):
    """Translate a source formula into the new layout, preserving exactly which
    cells it referenced (so totals keep the source's own inclusions/omissions)."""
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    def sub(m):
        d1, col, d2, row = m.groups()
        nc = colmap.get(col, col)
        nr = rowmap.get(int(row), int(row))
        return f'{d1}{nc}{d2}{nr}'

    return _REF.sub(sub, formula)


# ─────────────────── per-row reference photos ───────────────────
EMU = 9525   # EMU per pixel


def place_image(ws, col, row, path, w=120, h=105, dx=0, dy=3):
    """Anchor an image inside one cell, with a pixel offset so two photos can
    sit side by side in the same cell (as some source sheets do)."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    img = XLImage(path)
    img.width, img.height = w, h
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=col - 1, colOff=dx * EMU, row=row - 1, rowOff=dy * EMU),
        ext=XDRPositiveSize2D(w * EMU, h * EMU))
    ws.add_image(img)


def extract_media(src_path, outdir):
    """Dump a workbook's media and return {drawing_index: [(cell, filename), ...]}."""
    import os, re, zipfile
    from openpyxl.utils import get_column_letter
    os.makedirs(outdir, exist_ok=True)
    z = zipfile.ZipFile(src_path)
    for n in z.namelist():
        if n.startswith('xl/media/'):
            open(os.path.join(outdir, os.path.basename(n)), 'wb').write(z.read(n))
    sheet_to_drawing = {}
    for i in range(1, 12):
        try:
            r = z.read(f'xl/worksheets/_rels/sheet{i}.xml.rels').decode()
        except KeyError:
            continue
        m = re.search(r'drawing(\d+)\.xml', r)
        if m:
            sheet_to_drawing[i] = int(m.group(1))
    anchors = {}
    for dn in sorted(n for n in z.namelist() if re.match(r'xl/drawings/drawing\d+\.xml$', n)):
        idx = int(dn.split('drawing')[-1].split('.')[0])
        try:
            rels = z.read(f'xl/drawings/_rels/drawing{idx}.xml.rels').decode()
        except KeyError:
            rels = ''
        rm = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
        body = z.read(dn).decode()
        out = []
        for a in re.findall(r'<xdr:(?:two|one)CellAnchor[\s\S]*?</xdr:(?:two|one)CellAnchor>', body):
            f = re.search(r'<xdr:from><xdr:col>(\d+)</xdr:col>.*?<xdr:row>(\d+)</xdr:row>', a, re.S)
            rid = re.search(r'r:embed="(rId\d+)"', a)
            if f and rid:
                out.append((f'{get_column_letter(int(f.group(1)) + 1)}{int(f.group(2)) + 1}',
                            os.path.basename(rm[rid.group(1)])))
        anchors[idx] = out
    return anchors, sheet_to_drawing
