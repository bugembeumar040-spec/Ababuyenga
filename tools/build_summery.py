"""Reformat Washroom_Summery.xlsx (12 sheets) into FAE master format."""
import os
import zipfile
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from fmt import render, remap, merge_runs

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'
SRC = U + '78c4eb92-Washroom_Summery.xlsx'

os.makedirs('ws_imgs', exist_ok=True)
z = zipfile.ZipFile(SRC)
open('ws_imgs/urinal.jpeg', 'wb').write(z.read('xl/media/image1.jpeg'))

srcwb = openpyxl.load_workbook(SRC)
wb = openpyxl.Workbook()
wb.remove(wb.active)

W = dict(area=30.45, no=14.45, gen=20.45, asset=18.45, rem=30.45)


def wash(src, title, hdr_row, first, last, tot_rows, loc_col, gen_col, asset_cols,
         rem_col=None, loc_hdr='Area', gen_hdr='Washroom (Gender)', num_col=None,
         freeze='D3', photo_col=None):
    """Master layout: Area | Washroom no | Washroom (Gender) | assets... | Remarks"""
    plan = [(loc_hdr, loc_col, W['area']), ('Washroom no', num_col, W['no'])]
    if gen_col:
        plan.append((gen_hdr, gen_col, W['gen']))
    for c in asset_cols:
        plan.append((src.cell(row=hdr_row, column=c).value, c, W['asset']))
    plan.append(('Remarks', rem_col, W['rem']))

    cols = [(h, None, w) for h, _c, w in plan]
    colmap = {get_column_letter(sc): get_column_letter(i)
              for i, (_h, sc, _w) in enumerate(plan, start=1) if sc}
    off = 3 - first
    rowmap = {r: r + off for r in range(1, src.max_row + 3)}

    # resolve vertically-merged location labels, then number washrooms per block
    locs, cur = [], None
    for r in range(first, last + 1):
        v = src.cell(row=r, column=loc_col).value
        if v is not None:
            cur = v
        locs.append(cur)
    seq, prev, n = [], object(), 0
    for a in locs:
        n = 1 if a != prev else n + 1
        seq.append(n)
        prev = a

    grid = []
    for i, r in enumerate(list(range(first, last + 1)) + list(tot_rows)):
        row = []
        for j, (_h, sc, _w) in enumerate(plan):
            if sc is None and j == 1:                      # derived Washroom no
                row.append(seq[i] if i < len(locs) else None)
            elif sc is None:
                row.append(None)
            else:
                row.append(remap(src.cell(row=r, column=sc).value, colmap, rowmap))
        grid.append(row)

    ws = wb.create_sheet(title)
    totals = set(range(len(grid) - len(tot_rows), len(grid)))
    render(ws, cols, grid, style='washroom', freeze=freeze, total_rows=totals,
           label_span=2, photo_row=bool(photo_col), photo_cols={photo_col} if photo_col else ())
    merge_runs(ws, 1, 3, locs)
    return ws


# ── 1-2. Mall public + staff washrooms ──
wash(srcwb[' Washroom'], 'Mall Washroom Deatils', 2, 3, 62, [63, 64, 65],
     1, 2, [3, 4, 5, 6, 7], rem_col=8)
wash(srcwb['Staff Washroom '], 'Staff Washroom Deatils', 2, 3, 13, [14, 15, 16],
     1, 2, [3, 4, 5, 6, 7, 8])

# ── 3. External staff washrooms (source carries its own washroom number in col C) ──
ws = wash(srcwb['External Staff Washroom '], 'External Staff Washroom', 4, 5, 26, [27],
          1, 2, [4, 5, 6, 7, 8], rem_col=9, gen_hdr='Location', num_col=3)
ws['A25'], ws['C25'] = 'TOTAL ', None        # source put the label in the Location column

# ── 4-6. FAE washrooms ──
ws = wash(srcwb['FAE Washroom  '], 'FAE Washroom Details', 3, 4, 21, [22, 23, 24],
          1, 2, [3, 4, 5, 6], rem_col=7, photo_col=6)
im = XLImage('ws_imgs/urinal.jpeg')          # source anchored this on the Urinal header
im.width, im.height = 128, 140
ws.add_image(im, 'F2')

wash(srcwb[' FAE Handicap Washroom   '], 'FAE Handicap Washroom', 2, 3, 20, [21, 22, 23],
     1, 2, [3, 4, 5])
wash(srcwb['FAE Staff Washroom '], 'FAE Staff Washroom', 2, 3, 14, [15, 16, 17],
     1, 2, [3, 4, 5, 6, 7])

# ── 7. Mall handicap ──
ws = wash(srcwb[' Mall Handcup Washroom'], 'Mall Handicap Washroom', 2, 3, 10, [11, 12, 13],
          1, 2, [3, 4, 5])
ws.merge_cells('D5:F5')      # source C5:E5 - one value spanning the asset columns
ws.merge_cells('D8:F8')      # source C8:E8

# ── 8-9. Zabeel: two stacked blocks in the source ──
zb = srcwb['Zabeel ']
wash(zb, 'Zabeel Washroom Details', 2, 3, 16, [17], 3, 4, [5, 6, 7, 8, 9, 10], rem_col=11)
wash(zb, 'Zabeel Handicap Washroom', 20, 21, 32, [33], 3, 4, [5, 6, 7], rem_col=8)

# ── 10-11. China Town: two stacked blocks ──
ct = srcwb['CT-WR']
wash(ct, 'CT Washroom Details', 2, 3, 8, [9, 10, 11], 2, 3, [4, 5, 6, 7, 8])
wash(ct, 'CT Handicap Washroom', 14, 15, 18, [19, 20, 21], 2, 3, [4, 5, 6, 7])

# ── 12-13. Fountain Views: two side-by-side blocks ──
fv = srcwb['FV-WR']
wash(fv, 'FV Washroom Details', 2, 3, 24, [25, 26, 27], 2, 3, [4, 5, 6, 7, 8], rem_col=9)
wash(fv, 'FV Handicap Washroom', 2, 3, 22, [23, 24, 25], 11, 12, [13, 14, 15, 16])

# ── 14. Baby rooms: five property blocks in the source, stacked under a Property column ──
br = srcwb['Baby Room ']
cols = [('Property', None, 20.45), ('Area', None, 30.45), ('Washroom no', None, W['no']),
        ('Washroom (Gender)', None, W['gen']), ('Washbasin ', None, W['asset']),
        ('Tissue Dispenser', None, W['asset']), ('Remarks', None, W['rem'])]

BLOCKS = [   # property, loc_col, gen_col, washbasin, tissue, remarks, first, last, totals
    ('Malls',  1, 2, 3, 4, 5, 3, 24, [25]),
    ('FAE',    7, 8, 9, None, 10, 3, 18, [19, 20, 21]),
    ('Zabeel', 12, 13, 14, None, 15, 3, 6, []),
    ('CT',     12, 13, 14, None, 15, 10, 13, [14, 15, 16]),
    ('FV',     1, 2, 3, None, 5, 29, 48, [49, 50, 51]),
]

grid, totals, locs_all = [], set(), []
for prop, lc, gc, wc, tc, rc, first, last, tots in BLOCKS:
    off = 3 + len(grid) - first
    rowmap = {r: r + off for r in range(1, br.max_row + 3)}
    colmap = {get_column_letter(lc): 'B', get_column_letter(gc): 'D',
              get_column_letter(wc): 'E', get_column_letter(rc): 'G'}
    if tc:
        colmap[get_column_letter(tc)] = 'F'
    locs, cur = [], None
    for r in range(first, last + 1):
        v = br.cell(row=r, column=lc).value
        if v is not None:
            cur = v
        locs.append(cur)
    seq, prev, n = [], object(), 0
    for a in locs:
        n = 1 if a != prev else n + 1
        seq.append(n)
        prev = a
    for i, r in enumerate(range(first, last + 1)):
        grid.append([prop, locs[i], seq[i], br.cell(row=r, column=gc).value,
                     remap(br.cell(row=r, column=wc).value, colmap, rowmap),
                     remap(br.cell(row=r, column=tc).value, colmap, rowmap) if tc else None,
                     remap(br.cell(row=r, column=rc).value, colmap, rowmap)])
        locs_all.append(f'{prop}|{locs[i]}')
    for r in tots:
        totals.add(len(grid))
        grid.append([prop, remap(br.cell(row=r, column=lc).value, colmap, rowmap), None,
                     remap(br.cell(row=r, column=gc).value, colmap, rowmap),
                     remap(br.cell(row=r, column=wc).value, colmap, rowmap),
                     remap(br.cell(row=r, column=tc).value, colmap, rowmap) if tc else None,
                     remap(br.cell(row=r, column=rc).value, colmap, rowmap)])
        locs_all.append(f'__tot{len(grid)}')

ws = wb.create_sheet('Baby Room Details')
render(ws, cols, grid, style='washroom', freeze='E3', total_rows=totals, label_span=3)
merge_runs(ws, 2, 3, [l.split('|', 1)[-1] if '|' in l else l for l in locs_all])

# ── 15. Cross-property summary ──
src = srcwb['Summery ']
cols = [('Location ', None, 34.45)] + [(src.cell(row=3, column=c).value, None, 18.45)
                                       for c in range(4, 10)] + [('Remarks', None, 26.45)]
colmap = {get_column_letter(c): get_column_letter(c - 2) for c in range(3, 10)}
rowmap = {r: r - 1 for r in range(1, 30)}
grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in range(3, 10)] + [None]
        for r in range(4, 17)]
ws = wb.create_sheet('Washroom Assets Summery')
render(ws, cols, grid, style='compact', freeze='B3', total_rows={12}, label_span=1)

wb.save(OUT + 'Washroom_Summary_Details_2026.xlsx')
print('wrote Washroom_Summary_Details_2026.xlsx')
print(len(wb.sheetnames), 'tabs:', wb.sheetnames)
