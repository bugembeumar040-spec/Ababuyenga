"""Reformat the two bin inventories into FAE master format."""
import openpyxl
from openpyxl.styles import Font, Alignment
from fmt import render, remap, finalize, HDR_FILL, TOT_FILL

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'

# ───────────────────────── 1. Common area bins ─────────────────────────
src = openpyxl.load_workbook(U + '73296bca-Common_area__Bin.xlsx')['TDM- Common area Bins']

cols = [('SL', None, 10.45), ('Area', None, 25.45), ('Sub Area', None, 30.45),
        ('Qty', None, 15.45), ('Remarks', None, 30.45)]

colmap = {'B': 'A', 'C': 'B', 'D': 'C', 'E': 'D'}
rowmap = {r: r - 3 for r in range(6, 19)}          # src 6..18 -> out 3..15

grid = []
for r in range(6, 19):
    grid.append([remap(src.cell(row=r, column=c).value, colmap, rowmap)
                 for c in range(2, 6)] + [None])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Common Area Bins Details'
render(ws, cols, grid, style='compact', freeze='A3',
       total_rows={10, 11, 12}, label_span=3)

# identity merges carried over from the source
ws.merge_cells('B7:C7')      # Metro link bridge  (src C10:D10)
ws.merge_cells('B8:C8')      # Grand Parking      (src C11:D11)
ws.merge_cells('B9:B12')     # Lobby              (src C12:C15)
finalize(wb)
wb.save(OUT + 'TDM_Common_Area_Bins_Details_2026.xlsx')
print('wrote TDM_Common_Area_Bins_Details_2026.xlsx')

# ───────────────────────── 2. Steel bins ─────────────────────────
src = openpyxl.load_workbook(U + '451b2fc1-Steel_Bins.xlsx')['2025']

floors = [src.cell(row=2, column=c).value for c in range(3, 13)]   # C2..L2
cols = ([('Location', None, 34.45)]
        + [('Floor', f, 16.45) for f in floors]
        + [('Total', None, 14.45), ('Remarks', None, 26.45)])

colmap = {chr(ord('B') + i): chr(ord('A') + i) for i in range(12)}  # B..M -> A..L
rowmap = {r: r for r in range(1, 34)}                               # rows unchanged

grid = []
for r in range(3, 28):
    grid.append([remap(src.cell(row=r, column=c).value, colmap, rowmap)
                 for c in range(2, 14)] + [None])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Steel Bins Details 2025'
render(ws, cols, grid, style='compact', freeze='A3',
       total_rows={24}, label_span=1)

# stock reconciliation block, preserved from source rows 28-32
summary = [('Total no of steel Bin on the floor', src['B28'].value),
           ('Spare steel bin Available', src['B29'].value),
           ('Used & damaged stock for Disposed ', src['B30'].value),
           ('Bin Transfer ', src['B31'].value),
           ('Total  Steel bins', '=SUM(D29+D30+D31)')]
for i, (label, val) in enumerate(summary):
    r = 29 + i
    ws.row_dimensions[r].height = 25.0
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    v = ws.cell(row=r, column=4, value=val)
    v.font = Font(name='Calibri', size=14, bold=True, color='000000')
    v.fill = TOT_FILL
    v.alignment = Alignment(horizontal='center', vertical='center')

finalize(wb)
wb.save(OUT + 'TDM_Steel_Bins_Details_2026.xlsx')
print('wrote TDM_Steel_Bins_Details_2026.xlsx')
