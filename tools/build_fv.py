"""Fountain Views source files -> FAE master format.

Four unique sources (Fountain_Views_Assets and _2 are byte-identical):
  FVWashroom.xlsx          washroom / handicap / baby room / prayer room
  FV_Common_area_bins.xlsx round bins by tower
  FVSeating_area.xlsx      seating area, active and superseded
  Fountain_Views_Assets    asset list and planter pots
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from fmt import (render, remap, merge_runs, place_image, extract_media, finalize,
                 HDR_FILL)

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'
A = 18.45
wb = openpyxl.Workbook()
wb.remove(wb.active)
strays = []          # images that fall outside a tab's photo strip


def clean(v):
    """Drop broken-image artefacts; '#VALUE!' is not asset data."""
    return None if isinstance(v, str) and v.strip() == '#VALUE!' else v


# ═══════════ 1. FVWashroom: four sheets, header / photo row / data ═══════════
SRCF = U + 'e480f288-FVWashroom.xlsx'
anch, s2d = extract_media(SRCF, 'fv_wash')
swb = openpyxl.load_workbook(SRCF)


def fv_wash(sheet, title, ncol, first, last, tots, drawing):
    """Area | Washroom no | Washroom (Gender) | assets | Remarks, photo strip row 2."""
    src = swb[sheet]
    assets = [src.cell(row=2, column=c).value for c in range(3, ncol + 1)]
    cols = ([('Area', None, 30.45), ('Washroom no', None, 14.45),
             ('Washroom (Gender)', None, 20.45)]
            + [(a, None, A) for a in assets] + [('Remarks', None, 26.45)])
    colmap = {'A': 'A', 'B': 'C'}
    for c in range(3, ncol + 1):
        colmap[get_column_letter(c)] = get_column_letter(c + 1)
    rowmap = {r: r - first + 3 for r in range(1, src.max_row + 3)}

    locs, cur = [], None
    for r in range(first, last + 1):
        v = src.cell(row=r, column=1).value
        if v is not None:
            cur = v
        locs.append(cur)
    seq, prev, n = [], object(), 0
    for a in locs:
        n = 1 if a != prev else n + 1
        seq.append(n)
        prev = a

    grid = []
    for i, r in enumerate(list(range(first, last + 1)) + list(tots)):
        lead = ([locs[i], seq[i], src.cell(row=r, column=2).value] if i < len(locs)
                else [src.cell(row=r, column=1).value, None, src.cell(row=r, column=2).value])
        grid.append(lead + [clean(remap(src.cell(row=r, column=c).value, colmap, rowmap))
                            for c in range(3, ncol + 1)] + [None])

    ws = wb.create_sheet(title)
    photo_cols = set(range(4, ncol + 2))
    render(ws, cols, grid, style='compact', freeze='D3', photo_row=True,
           photo_cols=photo_cols, total_rows=set(range(len(grid) - len(tots), len(grid))),
           label_span=2, data_height=20.15)
    merge_runs(ws, 1, 3, locs)

    # row 3 of the source is its photo strip; keep genuine notes, drop '#VALUE!'
    for c in range(3, ncol + 1):
        v = clean(src.cell(row=3, column=c).value)
        if v is not None:
            cell = ws.cell(row=2, column=c + 1, value=v)
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cellref, img in anch[drawing]:
        col = openpyxl.utils.column_index_from_string(
            ''.join(ch for ch in cellref if ch.isalpha()))
        row = int(''.join(ch for ch in cellref if ch.isdigit()))
        tgt = col + 1 if col >= 3 else None
        if tgt and tgt in photo_cols and row <= 3:
            place_image(ws, tgt, 2, f'fv_wash/{img}', w=118, h=95)
        else:
            strays.append((title, cellref, f'fv_wash/{img}'))
    return ws


fv_wash('Washroom', 'FV Washroom (Detailed)', 23, 4, 25, [26, 27, 28], s2d[1])
fv_wash('Handicap Washroom   ', 'FV Handicap (Detailed)', 13, 4, 23, [24, 25, 26], s2d[2])
fv_wash('Baby Room ', 'FV Baby Room (Detailed)', 12, 4, 23, [24, 25, 26], s2d[3])
fv_wash('Prayer room ', 'FV Prayer Room', 11, 4, 5, [6], s2d[4])

# ═══════════ 2. Common area bins ═══════════
SRCF = U + '9c936189-FV_Common_area_bins.xlsx'
anch, s2d = extract_media(SRCF, 'fv_bins')
src = openpyxl.load_workbook(SRCF)['FV']
# column A is empty in every row and its header carries another property's name; dropped
cols = [(src.cell(row=2, column=c).value, None, w) for c, w in
        ((2, 28.45), (3, A), (4, 28.45), (5, A))] + [('Remarks', None, 24.45)]
colmap = {get_column_letter(c): get_column_letter(c - 1) for c in range(2, 6)}
rowmap = {r: r for r in range(1, 25)}
grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in range(2, 6)] + [None]
        for r in range(3, 18)]
ws = wb.create_sheet('FV Common Area Bins')
render(ws, cols, grid, style='compact', freeze='B3', total_rows={13, 14}, label_span=1)
for cellref, img in anch[s2d[1]]:
    strays.append(('FV Common Area Bins', cellref, f'fv_bins/{img}'))

# ═══════════ 3. Seating area ═══════════
SRCF = U + '4e0d7d24-FVSeating_area.xlsx'
anch, s2d = extract_media(SRCF, 'fv_seat')
swb = openpyxl.load_workbook(SRCF)


def seating(sheet, title, ncol, drawing):
    src = swb[sheet]
    heads = [src.cell(row=2, column=c).value for c in range(1, ncol + 1)]
    cols = [(heads[0], None, 10.45), (heads[1], None, 34.45), (heads[2], None, 22.45)] + \
           [(h, None, A) for h in heads[3:-1]] + [(heads[-1], None, 26.45)]
    colmap = {get_column_letter(c): get_column_letter(c) for c in range(1, ncol + 1)}
    rowmap = {r: r for r in range(1, 20)}
    grid = [[clean(remap(src.cell(row=r, column=c).value, colmap, rowmap))
             for c in range(1, ncol + 1)] for r in range(3, 8)]
    ws = wb.create_sheet(title)
    render(ws, cols, grid, style='compact', freeze='D3', total_rows={4},
           label_span=2, data_height=99.95)
    seen = {}
    for cellref, img in anch[drawing]:
        row = max(int(''.join(ch for ch in cellref if ch.isdigit())), 3)
        n = seen.get(row, 0); seen[row] = n + 1
        place_image(ws, 3, row, f'fv_seat/{img}', w=64, h=92, dx=2 + n * 68)


seating('FV New- active', 'FV Seating Area New', 12, s2d[1])
seating('FV Old', 'FV Seating Area Old', 11, s2d[2])

# ═══════════ 4. Assets and planter pots ═══════════
SRCF = U + '90238346-Fountain_Views_Assets.xlsx'
anch, s2d = extract_media(SRCF, 'fv_assets')
swb = openpyxl.load_workbook(SRCF)

src = swb['Assets']
cols = [('No', None, 10.45), ('Photo', None, 22.45), ('Description', None, 40.45),
        ('Quantity on Floor', None, A), ('Remarks', None, 24.45)]
grid, photo_src = [], []
for c0, first, last in ((1, 3, 25), (7, 3, 20)):        # two side-by-side blocks -> stacked
    for r in range(first, last + 1):
        vals = [clean(src.cell(row=r, column=c0 + i).value) for i in range(5)]
        if all(v is None for v in vals):
            continue
        grid.append(vals)
        photo_src.append((get_column_letter(c0 + 1), r))
ws = wb.create_sheet('FV Assets')
render(ws, cols, grid, style='compact', freeze='C3', data_height=109.5)
lookup = {c: f for c, f in anch[s2d[1]]}
for i, (col, r) in enumerate(photo_src):
    f = lookup.get(f'{col}{r}')
    if f:
        place_image(ws, 2, 3 + i, f'fv_assets/{f}', w=128, h=120)

src = swb['FV Planter pot']
grid = [[clean(src.cell(row=r, column=c).value) for c in range(1, 6)] for r in range(3, 12)]
ws = wb.create_sheet('FV Planter Pots')
render(ws, cols, grid, style='compact', freeze='C3', data_height=109.5)
lookup = {c: f for c, f in anch[s2d[2]]}
for i, r in enumerate(range(3, 12)):
    f = lookup.get(f'B{r}')
    if f:
        place_image(ws, 2, 3 + i, f'fv_assets/{f}', w=128, h=120)

# ═══════════ 5. Photo reference for images outside a strip ═══════════
if strays:
    ws = wb.create_sheet('FV Photo Reference')
    cols = [('Source tab', None, 28.45), ('Shown at', None, 18.45),
            ('Photo', None, 22.45), ('Remarks', None, 24.45)]
    render(ws, cols, [[t, c, None, None] for t, c, _f in strays],
           style='compact', freeze='C3', data_height=115.0)
    for i, (_t, _c, f) in enumerate(strays):
        place_image(ws, 3, 3 + i, f, w=128, h=145)

finalize(wb)
wb.save(OUT + 'FV_Source_Assets_Details_2026.xlsx')
print('wrote FV_Source_Assets_Details_2026.xlsx')
print(f'  {len(wb.sheetnames)} tabs, {sum(len(w._images) for w in wb.worksheets)} photos')
print(' ', wb.sheetnames)
