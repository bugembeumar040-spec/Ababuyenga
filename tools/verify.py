"""Verify the reformatted workbooks reproduce the source's computed numbers.

LibreOffice cannot run in this container, so instead of recalculating we
evaluate every formula in the output and compare the resulting multiset of
values against the values Excel cached in the source workbook.
"""
import re
from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'

REF = re.compile(r'\b([A-Z]{1,3})(\d+)\b')
RANGE = re.compile(r'\b([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)\b')


def evaluate_sheet(ws):
    """Iteratively evaluate every formula cell in a sheet."""
    vals, formulas = {}, {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                formulas[c.coordinate] = c.value
            elif isinstance(c.value, (int, float)):
                vals[c.coordinate] = c.value
            elif c.value is not None:
                vals[c.coordinate] = c.value

    def expand(expr):
        def rng(m):
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            a, b = column_index_from_string(c1), column_index_from_string(c2)
            cells = [f'{get_column_letter(col)}{r}'
                     for col in range(min(a, b), max(a, b) + 1)
                     for r in range(min(r1, r2), max(r1, r2) + 1)]
            return '(' + '+'.join(cells) + ')' if cells else '(0)'
        return RANGE.sub(rng, expr)

    for _ in range(12):                     # fixpoint: formulas may cite formulas
        progress = False
        for coord, f in list(formulas.items()):
            if coord in vals:
                continue
            e = f[1:]
            # SUBTOTAL(9, range) is a plain SUM for our purposes
            e = re.sub(r'SUBTOTAL\(\s*9\s*,', 'SUM(', e, flags=re.I)
            counta = e.upper().startswith('COUNTA')
            if counta:
                inner = e[e.index('(') + 1:e.rindex(')')]
                m = RANGE.search(inner)
                if not m:
                    continue
                c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                a, b = column_index_from_string(c1), column_index_from_string(c2)
                n = 0
                for col in range(min(a, b), max(a, b) + 1):
                    for r in range(min(r1, r2), max(r1, r2) + 1):
                        cd = f'{get_column_letter(col)}{r}'
                        if vals.get(cd) not in (None, '') or cd in formulas:
                            n += 1
                vals[coord] = n
                progress = True
                continue
            e = expand(e)
            e = e.replace(',', '+').replace('SUM', '')
            missing = False

            def sub(m):
                nonlocal missing
                cd = m.group(0)
                if cd in vals:
                    v = vals[cd]
                    return str(v) if isinstance(v, (int, float)) else '0'
                if cd in formulas:
                    missing = True
                    return '0'
                return '0'
            e2 = REF.sub(sub, e)
            if missing:
                continue
            try:
                vals[coord] = eval(e2)      # arithmetic over resolved cell values
                progress = True
            except Exception:
                vals[coord] = f'ERR:{f}'
                progress = True
        if not progress:
            break
    return {k: vals.get(k, f'UNRESOLVED:{v}') for k, v in formulas.items()}


def cached(path, sheet):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    out = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, (int, float)):
                out.append(c.value)
    return out


def formula_cached(src_path, sheet):
    """Values Excel cached for the source sheet's own formula cells."""
    f = openpyxl.load_workbook(src_path)[sheet]
    v = openpyxl.load_workbook(src_path, data_only=True)[sheet]
    res = []
    for row in f.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                res.append(v[c.coordinate].value)
    return res


CASES = [
    ('TDM_Common_Area_Bins_Details_2026.xlsx', ['Common Area Bins Details'],
     '73296bca-Common_area__Bin.xlsx', ['TDM- Common area Bins']),
    ('TDM_Steel_Bins_Details_2026.xlsx', ['Steel Bins Details 2025'],
     '451b2fc1-Steel_Bins.xlsx', ['2025']),
    ('TDM_Mall_Washroom_Details_2026.xlsx',
     ['Mall Washroom Deatils', 'Staff Washroom Deatils', 'Baby Room Details',
      'Handicap Washroom Details', 'Washroom Assets Summery'],
     'a62f1923-Mall_Washroom.xlsx',
     [' Washroom', 'Staff Washroom ', 'Baby Room', 'Handcup Washroom', 'Summery ']),
    ('Planters_in_Malls_Details_2026.xlsx',
     ['RC & IR Lobbies', 'FP Lobbies', 'Mall Planters'],
     '43ae292f-Planters_in_Malls.xlsx', ['RC & IR Lobbies', 'FP Lobbies', 'Mall']),
    ('Washroom_Summary_Details_2026.xlsx',
     ['Mall Washroom Deatils', 'Staff Washroom Deatils', 'External Staff Washroom',
      'FAE Washroom Details', 'FAE Handicap Washroom', 'FAE Staff Washroom',
      'Mall Handicap Washroom', 'Zabeel Washroom Details', 'Zabeel Handicap Washroom',
      'CT Washroom Details', 'CT Handicap Washroom', 'FV Washroom Details',
      'FV Handicap Washroom', 'Baby Room Details', 'Washroom Assets Summery'],
     '78c4eb92-Washroom_Summery.xlsx',
     [' Washroom', 'Staff Washroom ', 'External Staff Washroom ', 'FAE Washroom  ',
      ' FAE Handicap Washroom   ', 'FAE Staff Washroom ', ' Mall Handcup Washroom',
      'Zabeel ', 'Zabeel ', 'CT-WR', 'CT-WR', 'FV-WR', 'FV-WR', 'Baby Room ', 'Summery ']),
]

overall_ok = True
for outfile, outsheets, srcfile, srcsheets in CASES:
    wb = openpyxl.load_workbook(OUT + outfile)
    got, errs = [], []
    for s in outsheets:
        res = evaluate_sheet(wb[s])
        for coord, v in res.items():
            if isinstance(v, str):
                errs.append(f'{s}!{coord} {v}')
            else:
                got.append(round(v, 6))
    want = []
    for s in set(srcsheets):
        want += [round(v, 6) for v in formula_cached(U + srcfile, s) if isinstance(v, (int, float))]

    gc, wc = Counter(got), Counter(want)
    only_out, only_src = gc - wc, wc - gc
    ok = not errs and not only_out and not only_src
    overall_ok &= ok
    print(f'{"OK  " if ok else "DIFF"} {outfile}')
    print(f'      formulas evaluated: {len(got)}   source cached: {len(want)}')
    if errs:
        print('      ERRORS:', errs[:8])
    if only_out or only_src:
        print('      only in output:', sorted(only_out.elements())[:15])
        print('      only in source:', sorted(only_src.elements())[:15])

print('\nALL MATCH' if overall_ok else '\nDIFFERENCES FOUND')
