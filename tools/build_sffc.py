"""Batch 2, file 5: SF_FC.xlsx -> FAE master format (5 tabs)."""
import openpyxl
from openpyxl.utils import get_column_letter
from fmt import render, remap, merge_runs, HDR_FILL, TOT_FILL
from openpyxl.styles import Font, Alignment

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'
A = 18.45

swb = openpyxl.load_workbook(U + '7738a4b2-SF_FC.xlsx')
wb = openpyxl.Workbook(); wb.remove(wb.active)

# ═════════ External: source stores each terrace as a label row + a value row ═════════
src = swb['External']
PAIRS = [(11, 12), (13, 14), (15, 16), (17, 18), (19, 20), (21, 22)]
items, recs = [], []
for lrow, vrow in PAIRS:
    t = str(src.cell(row=lrow, column=1).value).replace('\xa0', ' ').strip()
    cat, loc = t[-6:], t[:-6].strip()          # trailing 'Chairs' / 'TABLES', source casing kept
    vals, variants = {}, []
    for c in (2, 3):
        lbl = src.cell(row=lrow, column=c).value
        val = src.cell(row=vrow, column=c).value
        if lbl is None:
            continue
        lbl = lbl.strip()
        match = next((i for i in items if i.lower() == lbl.lower()), None)
        if match is None:
            items.append(lbl); match = lbl
        elif match != lbl:
            variants.append(f'source spelled it "{lbl}"')
        vals[match] = val
    recs.append((loc, cat, vals, src.cell(row=vrow, column=7).value,
                 '; '.join(variants) or None))

cols = ([('Location', None, 24.45), ('Category', None, 14.45)]
        + [(i, None, A) for i in items]
        + [('Total On the floor', None, A), ('Remarks', None, 30.45)])
grid = [[loc, cat] + [vals.get(i) for i in items] + [tot, rem]
        for loc, cat, vals, tot, rem in recs]
ws = wb.create_sheet('SF FC External')
render(ws, cols, grid, style='compact', freeze='C3')
merge_runs(ws, 1, 3, [g[0] for g in grid])

# preserve the source's superseded header block (rows 1-5) rather than drop it
r0 = 3 + len(grid) + 1
lbl = ws.cell(row=r0, column=1, value='Source header block (rows 1-5, superseded)')
lbl.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
lbl.fill = HDR_FILL
lbl.alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=len(cols))
for i, sr in enumerate((1, 2, 4, 5)):
    rr = r0 + 1 + i
    ws.row_dimensions[rr].height = 25.0
    for j, c in enumerate((1, 3, 4, 5, 7), start=1):
        cell = ws.cell(row=rr, column=j, value=src.cell(row=sr, column=c).value)
        cell.font = Font(name='Calibri', size=12, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ═════════ Internal: chairs block and tables block sit side by side ═════════
src = swb['Internal ']


def block(title, c0, ncols, subtotal_col, freeze='B3'):
    heads = [src.cell(row=3, column=c).value for c in range(c0, c0 + ncols)]
    cols = ([(heads[0], None, 26.45)] + [(h, None, A) for h in heads[1:]]
            + [('Remarks', None, 34.45)])
    colmap = {get_column_letter(c0 + i): get_column_letter(1 + i) for i in range(ncols)}
    rowmap = {r: r - 1 for r in range(1, src.max_row + 3)}
    grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap)
             for c in range(c0, c0 + ncols)] + [None] for r in range(4, 20)]
    ws = wb.create_sheet(title)
    render(ws, cols, grid, style='compact', freeze=freeze,
           total_rows={13, 14, 15}, label_span=1)
    return ws, colmap


block('SFFC Chairs In Use', 1, 11, 'K')
ws, _ = block('SFFC Tables In Use', 12, 7, 'R')
# source R18 reads =SUM(I18:P18) - columns that live in the chairs block, not here.
# Both it and the in-block equivalent evaluate to 0 (row 18 is empty); noted, not silently changed.
ws['G17'] = '=SUM(B17:F17)'
ws['H17'] = 'Source formula was =SUM(I18:P18), which spans columns outside the tables block; both evaluate to 0'

# leather sofa block
cols = [('S/N', None, 10.45), ('LOCATION', None, 20.45), ('Length in cm', None, A),
        ('Width in cm', None, A), ('qty', None, A), ('Remarks', None, 24.45)]
colmap = {get_column_letter(2 + i): get_column_letter(1 + i) for i in range(5)}
rowmap = {r: r - 22 for r in range(1, src.max_row + 3)}
grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap)
         for c in range(2, 7)] + [None] for r in range(25, 30)]
ws = wb.create_sheet('SFFC Leather Sofa')
render(ws, cols, grid, style='compact', freeze='C3', total_rows={4}, label_span=2)

# ═════════ Summery ═════════
src = swb['Summery ']
cols = ([('Location', None, 30.45)]
        + [(src.cell(row=4, column=c).value, None, A) for c in range(4, 10)]
        + [('Remarks', None, 24.45)])
colmap = {get_column_letter(3 + i): get_column_letter(1 + i) for i in range(7)}
rowmap = {r: r - 2 for r in range(1, 20)}
grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap)
         for c in range(3, 10)] + [None] for r in range(5, 8)]
ws = wb.create_sheet('SF FC Summery')
render(ws, cols, grid, style='compact', freeze='B3')
for rng in ('D3:D4', 'E3:E5', 'F3:F5', 'G3:G5'):     # source F5:F6, G5:G7, H5:H7, I5:I7
    ws.merge_cells(rng)

wb.save(OUT + 'SF_FC_Furniture_Details_2026.xlsx')
print('SF_FC_Furniture_Details_2026.xlsx', wb.sheetnames)
print('asset columns found:', items)
