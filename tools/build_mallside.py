"""Mall side Assets Details 2026 -> FAE master format.

This source is already built on the master's own template (same 64-column
washroom layout, same prayer-room columns, same photo-strip idea), so the job is
mostly to re-render it through the shared formatter and normalise the
common-area sheet, which the source splits into four stacked per-floor blocks.
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter
from fmt import render, remap, merge_runs, place_image, extract_media, finalize

U = '/root/.claude/uploads/342efb5e-df1f-5d88-ab0b-269e90eb1eb9/'
OUT = '/home/user/Ababuyenga/reports/'
SRCF = U + 'c4e99282-Mall_side__Assets_Details_2026.xlsx'
IMG = 'ms_imgs'

anchors, sheet2draw = extract_media(SRCF, IMG)
swb = openpyxl.load_workbook(SRCF)
wb = openpyxl.Workbook()
wb.remove(wb.active)

WASH = 'Mall Side Washroom Deatils'
PRAY = 'Mall Side Prayer Room'
COMM = 'Mall Side Common Area'


def read_header(ws, ncol, hdr1=1, hdr2=2):
    """Recover (group, sub) per column from a master-format two-row header."""
    vmerge, hmerge = {}, {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row == hdr1 and rng.max_row == hdr2:          # X1:X2 -> flat
            for c in range(rng.min_col, rng.max_col + 1):
                vmerge[c] = ws.cell(row=hdr1, column=rng.min_col).value
        elif rng.min_row == rng.max_row == hdr1:                  # group band
            for c in range(rng.min_col, rng.max_col + 1):
                hmerge[c] = ws.cell(row=hdr1, column=rng.min_col).value
    out = []
    for c in range(1, ncol + 1):
        if c in vmerge:
            out.append((vmerge[c], None))
        else:
            g = hmerge.get(c, ws.cell(row=hdr1, column=c).value)
            out.append((g, ws.cell(row=hdr2, column=c).value))
    return out


# ═══════════════ 1. Washroom (already the master's 64-column layout) ═══════════════
src = swb['Mall side Washroom Deatils']
hdr = read_header(src, 64)
W = {1: 13.45, 2: 40.73, 3: 46.45, 64: 42.54}
cols = [(g, s, W.get(i, 15.45)) for i, (g, s) in enumerate(hdr, start=1)]

rowmap = {r: r for r in range(1, 80)}          # data starts at row 3 in both
colmap = {}
areas, cur = [], None
for r in range(3, 68):
    v = src.cell(row=r, column=1).value
    if v is not None:
        cur = v
    areas.append(cur)

grid = [[remap(src.cell(row=r, column=c).value, colmap, rowmap) for c in range(1, 65)]
        for r in range(3, 70)]                 # keep the full span incl. blank row 68
ws = wb.create_sheet(WASH)
render(ws, cols, grid, style='washroom', freeze='D3',
       total_rows={69 - 3}, label_span=3)
merge_runs(ws, 1, 3, areas)

# ═══════════════ 2. Prayer room ═══════════════
src = swb['Prayer Room Details']
hdr = read_header(src, 23)
W = {1: 13.45, 2: 13.54, 3: 42.54, 23: 30.0}
cols = [(g, s, W.get(i, 15.45)) for i, (g, s) in enumerate(hdr, start=1)]
areas, cur = [], None
for r in range(3, 19):
    v = src.cell(row=r, column=1).value
    if v is not None:
        cur = v
    areas.append(cur)
grid = []
for r in range(3, 23):
    row = []
    for c in range(1, 24):
        v = src.cell(row=r, column=c).value
        if isinstance(v, str):                 # retarget the cross-sheet reference
            v = v.replace("'Mall side Washroom Deatils'", f"'{WASH}'")
        row.append(remap(v, {}, {x: x for x in range(1, 80)}))
    grid.append(row)
ws = wb.create_sheet(PRAY)
render(ws, cols, grid, style='washroom', freeze='D3',
       total_rows={22 - 3}, label_span=3)
merge_runs(ws, 1, 3, areas)

# ═══════════════ 3. Common area: four stacked per-floor blocks -> one table ═══════════════
src = swb['Common Area Assets Details']
# header row, photo row, data rows, total row. The data span runs right up to the
# total row (blank source rows included) so each block's SUM range stays inside it.
BLOCKS = [
    (1, 2, [3, 4], 5),
    (9, 10, [11, 12], 13),
    (17, 18, [19, 20, 21, 22, 23], 24),
    (27, 28, [29, 30, 31], 32),
]
photo_at = {c: f for c, f in anchors[sheet2draw[3]]}

union, seen_in_block = [], {}
block_maps = []
for hdr_row, _p, _d, _t in BLOCKS:
    last = max(c for c in range(3, 25) if src.cell(row=hdr_row, column=c).value is not None)
    names, mapping = {}, {}
    for c in range(3, last):                   # last column of a block is its Remarks
        name = src.cell(row=hdr_row, column=c).value
        n = names.get(name, 0) + 1
        names[name] = n
        key = name if n == 1 else f'{name} ({n})'
        if key not in union:
            union.append(key)
            seen_in_block[key] = (hdr_row, c)
        mapping[c] = key
    block_maps.append((mapping, last))

cols = ([('Level', None, 25.45), ('Location', None, 28.45)]
        + [(k, None, 18.54) for k in union] + [('Remarks', None, 17.82)])
col_of = {k: i for i, k in enumerate(union, start=3)}

grid, totals, levels = [], set(), []
for (hdr_row, photo_row, data_rows, total_row), (mapping, last) in zip(BLOCKS, block_maps):
    rowmap, colmap = {}, {}
    for i, sr in enumerate(data_rows):
        rowmap[sr] = 3 + len(grid) + i
    rowmap[total_row] = 3 + len(grid) + len(data_rows)
    for sc, key in mapping.items():
        colmap[get_column_letter(sc)] = get_column_letter(col_of[key])
    lvl = next((src.cell(row=r, column=1).value for r in data_rows
                if src.cell(row=r, column=1).value), None)
    loc = next((src.cell(row=r, column=2).value for r in data_rows
                if src.cell(row=r, column=2).value), None)
    for sr in data_rows:
        vals = {key: remap(src.cell(row=sr, column=sc).value, colmap, rowmap)
                for sc, key in mapping.items()}
        rem = src.cell(row=sr, column=last).value
        populated = any(v is not None for v in vals.values()) or rem is not None \
            or src.cell(row=sr, column=1).value is not None
        row = [lvl if populated else None, loc if populated else None] \
            + [None] * (len(union) + 1)
        for key, v in vals.items():
            row[col_of[key] - 1] = v
        row[-1] = rem
        grid.append(row)
        levels.append(lvl if populated else f'__blank{len(grid)}')
    trow = [src.cell(row=total_row, column=1).value, None] + [None] * (len(union) + 1)
    for sc, key in mapping.items():
        trow[col_of[key] - 1] = remap(src.cell(row=total_row, column=sc).value, colmap, rowmap)
    totals.add(len(grid))
    grid.append(trow)
    levels.append(f'__tot{len(grid)}')

ws = wb.create_sheet(COMM)
render(ws, cols, grid, style='compact', freeze='C3', photo_row=True,
       photo_cols=set(col_of.values()), total_rows=totals, label_span=2)
merge_runs(ws, 1, 3, levels)

placed = 0
for key, (hdr_row, sc) in seen_in_block.items():
    photo_row = next(p for h, p, _d, _t in BLOCKS if h == hdr_row)
    f = photo_at.get(f'{get_column_letter(sc)}{photo_row}')
    if f and f.lower().endswith(('.png', '.jpg', '.jpeg')):
        place_image(ws, col_of[key], 2, os.path.join(IMG, f), w=118, h=132)
        placed += 1

finalize(wb)
wb.save(OUT + 'Mall_Side_Assets_Details_2026.xlsx')
print('wrote Mall_Side_Assets_Details_2026.xlsx', wb.sheetnames)
print(f'  union asset columns: {len(union)}   photos placed: {placed}')
